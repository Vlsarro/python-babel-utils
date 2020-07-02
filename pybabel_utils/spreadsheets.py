# -*- coding: utf-8 -*-

import errno
import os

from babel.util import odict
from babel.messages.pofile import read_po, write_po

from pybabel_utils import logger
from pybabel_utils.catalog import UpdatableCatalog

try:
    import openpyxl
except ImportError:
    openpyxl = None


ENCODING = 'utf-8'


def get_po_filenames(dir_path):
    filenames = []
    for f in os.listdir(dir_path):
        if f.endswith('.po'):
            filenames.append(os.path.join(dir_path, f))
    return filenames


def get_po_filename_from_path(po_filepath):
    return po_filepath.split(os.sep)[-1]


class PoFilesSpreadsheetExporter(object):

    @classmethod
    def _init_po_workbook(cls, po_filenames):
        """Create workbook and fill header row"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        col = 1
        row = 1
        worksheet.cell(row=row, column=col, value='msgid')

        col += 1
        for name in po_filenames:
            worksheet.cell(row=row, column=col, value=get_po_filename_from_path(name))
            col += 1

        return workbook, worksheet

    @staticmethod
    def _rollback_obsolete(catalog):
        catalog._messages.update(catalog.obsolete)
        catalog.obsolete = odict()

    @classmethod
    def _fill_po_worksheet(cls, po_filenames, worksheet, include_obsolete):
        col = 1
        row = 2

        for filename in po_filenames:
            with open(filename) as f:
                cat = read_po(f, ignore_obsolete=not include_obsolete)

                if include_obsolete:
                    cls._rollback_obsolete(cat)

                for msg in cat:
                    if msg.id != '':  # ignore header
                        if col == 1:
                            worksheet.cell(row=row, column=col, value=msg.id)
                            worksheet.cell(row=row, column=col + 1, value=msg.string)
                        else:
                            worksheet.cell(row=row, column=col + 1, value=msg.string)
                        row += 1
            col += 1
            row = 2

    @classmethod
    def run(cls, input_folder_name, output_filename, include_commented_messages=False, save=True):
        if openpyxl is None:
            raise ImportError('Please install "openpyxl" to use "PoFilesSpreadsheetExporter" class')

        po_filenames = get_po_filenames(input_folder_name)
        if not po_filenames:
            raise ValueError('No files were added. Please, try again.')

        wb, ws = cls._init_po_workbook(po_filenames)
        try:
            cls._fill_po_worksheet(po_filenames, ws, include_commented_messages)
            if save:
                wb.save(filename=output_filename)
        except Exception as e:
            logger.debug('Exception occurred while worksheet processing > {!r}'.format(e), exc_info=True)

        return wb


class PoFilesSpreadsheetUpdater(object):

    @staticmethod
    def _get_column_index(row, po_file_short_name):
        for cell in row:
            if po_file_short_name in cell.value:
                return cell.column
            else:
                logger.debug('Column [{}] doesn\'t have data for {}'.format(cell.value, po_file_short_name))
        return None

    @classmethod
    def _update_po_files(cls, input_spreadsheet, po_filenames, output_dir, save):
        updated_catalogs = []

        if output_dir:
            try:
                os.makedirs(output_dir)
            except OSError as e:
                if e.errno != errno.EEXIST:
                    output_dir = None

        wb = openpyxl.load_workbook(input_spreadsheet, read_only=True)
        ws = wb.active

        first_row = ws[1]

        for filename in po_filenames:
            with open(filename) as f:
                short_filename = get_po_filename_from_path(filename)
                logger.debug('\n--------- Processing [{}] ---------\n'.format(short_filename))

                col_index = cls._get_column_index(first_row, short_filename)
                if col_index:
                    catalog = read_po(f, charset=ENCODING)
                    new_catalog = cls._process_catalog(ws, catalog, col_index)

                    updated_catalogs.append(new_catalog)

                    if output_dir:
                        file_args = (os.path.join(output_dir, short_filename), 'w+')
                    else:
                        file_args = (filename, 'w+')

                    if save:
                        with open(*file_args) as nf:
                            write_po(nf, new_catalog, width=None)

        return updated_catalogs

    @classmethod
    def _process_catalog(cls, worksheet, catalog, column_idx):
        new_catalog = UpdatableCatalog()
        new_catalog.__dict__ = catalog.__dict__.copy()

        for row in worksheet.iter_rows(min_row=2):
            if row:
                msgid = row[0].value
                msgstr = row[column_idx - 1].value
                if msgid and msgstr:
                    new_catalog.add(msgid, string=msgstr)
                    logger.debug(u'\nMsgid: {}\nMsgstr: {}\nMsgid: {}\nMsgstr (catalog): {}\n'.format(
                        msgid, msgstr, new_catalog[msgid], new_catalog[msgid].string)
                    )

        return new_catalog

    @classmethod
    def run(cls, input_spreadsheet, input_folder_name, output_dir=None, save=True):
        if openpyxl is None:
            raise ImportError('Please install "openpyxl" to use "PoFilesSpreadsheetUpdater" class')

        po_filenames = get_po_filenames(input_folder_name)
        if not po_filenames:
            raise ValueError('No files were added. Please, try again.')

        try:
            updated_catalogs = cls._update_po_files(input_spreadsheet, po_filenames, output_dir, save)
        except Exception as e:
            logger.debug('Exception occurred while po files update > {!r}'.format(e), exc_info=True)
        else:
            return updated_catalogs
